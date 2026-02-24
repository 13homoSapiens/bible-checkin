
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

def beauty():

    file_path = "result.xlsx"
    wb = load_workbook(file_path)
    ws = wb.active

    # 1️⃣ 全部单元格居中 + 楷体
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="楷体", size=12)

    # 2️⃣ 设置列宽
    ws.column_dimensions["A"].width = 14  # 姓名列宽
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 6
    for col in range(4, ws.max_column + 1):
        ws.column_dimensions[chr(64 + col)].width = 16

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    gray_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    ws.conditional_formatting.add(
        "B2:B100",
        FormulaRule(formula=["B2=7"], fill=orange_fill)
    )

    ws.conditional_formatting.add(
        "D2:Z100",
        FormulaRule(formula=['D2="(^_^)"'], fill=green_fill)
    )

    ws.conditional_formatting.add(
        "D2:Z100",
        FormulaRule(formula=['D2="(~~~)"'], fill=gray_fill)
    )

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in [None, ""]:
                cell.border = border

    wb.save(file_path)